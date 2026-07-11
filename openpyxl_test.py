from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl

wb = openpyxl.Workbook()

ws = wb.active

ws.title = "成绩表"

headers = ["姓名", "数学", "语文", "英语", "总分"]

ws.append(headers)

data = [

    ["小红", 90, 85, 92, ],

    ["小明", 85, 80, 88, ],

    ["小强", 92, 87, 94, ],

    ["小花", 88, 84, 91, ],

    ["张三", 76, 81, 79, ],

    ["李四", 95, 89, 97, ],

]

for row in data:

    ws.append(row)

wb.save("data.xlsx")

print("data.xlsx已成功创建!")


wb = openpyxl.load_workbook("data.xlsx")

ws = wb["成绩表"]

# 单元格读取

value = ws["A1"].value

# value = ws.cell(1, 1).value

print("A1单元格的内容:", value)

# 批量读取

for row in ws.iter_rows(values_only=True):

    for cell in row:

        print(cell, end="  ")

    print()

ws["A2"] = "赛博红兔"

wb.save("data.xlsx")

print("Excel数据已修改!")

# 遍历每行并计算总分

for row in range(2, ws.max_row + 1):

    ws[f"E{row}"] = f"=SUM(B{row}:D{row})"

wb.save("data.xlsx")

print("总分计算完毕!")


cell = ws["A1"]

cell.font = Font(name="Calibri", bold=True, italic=True,

                 color="FF0000", size=14)

cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("data.xlsx")

print("Excel格式化完成!")
